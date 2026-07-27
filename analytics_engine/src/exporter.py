"""Exporter — Data export in multiple formats."""

import csv
import json
import io
from datetime import datetime
from loguru import logger


class ExporterService:
    """Export analytics and campaign data in various formats."""

    def export_to_csv(self, data: list[dict], filename: str = "export.csv") -> str:
        """Export data as CSV."""
        if not data:
            return ""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        content = output.getvalue()
        output.close()
        logger.info(f"CSV exported: {len(data)} rows")
        return content

    def export_to_json(self, data: dict, filename: str = "export.json") -> str:
        """Export data as JSON."""
        content = json.dumps(data, indent=2, default=str)
        logger.info(f"JSON exported: {len(content)} bytes")
        return content

    def export_to_excel(self, sheets: dict[str, list[dict]]) -> bytes:
        """Export multiple sheets to Excel format."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            for sheet_name, data in sheets.items():
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                if data:
                    for row_idx, row in enumerate(data, 1):
                        for col_idx, (key, value) in enumerate(row.items(), 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            logger.info(f"Excel exported: {len(sheets)} sheets")
            return output.read()
        except ImportError:
            logger.warning("openpyxl not installed, falling back to JSON")
            return json.dumps({"sheets": {k: v for k, v in sheets.items()}, "_format": "json_fallback"}).encode()

    def export_api_endpoint(self, data: dict) -> dict:
        """Format data for REST API download endpoint."""
        return {
            "data": data,
            "exported_at": datetime.utcnow().isoformat(),
            "format": "json",
            "download_url": "/api/v1/analytics/export/download",
        }

    def export_campaign_report(self, campaign_id: str, report_data: dict, fmt: str = "json") -> str | bytes:
        """Full campaign report export."""
        wrapper = {
            "campaign_id": campaign_id,
            "exported_at": datetime.utcnow().isoformat(),
            "report": report_data,
        }

        if fmt == "csv":
            rows = []
            for key, value in report_data.items():
                if isinstance(value, (int, float, str)):
                    rows.append({"metric": key, "value": value})
            return self.export_to_csv(rows)
        elif fmt == "json":
            return self.export_to_json(wrapper)
        elif fmt == "excel":
            return self.export_to_excel({"summary": [wrapper], "details": report_data.get("details", [])})
        return json.dumps(wrapper, default=str)
